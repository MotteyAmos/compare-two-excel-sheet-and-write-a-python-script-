from openpyxl import load_workbook, Workbook
import re


def wasc_leg_name_report(workbook1, workbook2, reportSavedWorkbook):
    #extract worksheet from workbooks
    wac_sheet = load_workbook(filename=workbook1).active
    leg_sheet = load_workbook(filename=workbook2).active

    report_workbook = Workbook()
    # create a report worksheet
    report_sheet = report_workbook.create_sheet("report_sheet")


    # return the data of birth and name of student in an object of array
    # by extracting them from the sheet
    def getName_N_DOB(sheet):
        name_DOB = []
        for value in sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
            otherNames = value[1]
            surName = value[0]
            DOB = str(value[3]).split()
            DOB = re.sub(r'[^\w\s]', ' ', DOB[0]).split(" ")


            # since set data structure don't arrange element in an order
            # I am accessing a charater to sort out the difference i.e day, month and year
            new_DOB = ""
            symbol_vale = "ymd"
            i=0

            for dof in DOB:
                dof = symbol_vale[i] + dof
                i += 1
                new_DOB = "{} {}".format(new_DOB, dof)

            new_DOB_array = new_DOB.split()
            new_DOB_array.sort()


            surName = re.sub(r'[^\w\s]', ' ', surName)
            otherNames = re.sub(r'[^\w\s]', ' ', otherNames)

            names = surName.lower() + " " + otherNames.lower()

            person = {
                "name": set(names.split()),
                "dateOfBirth": set(new_DOB_array)
            }

            name_DOB.append(person)
        return name_DOB


    wac_names_DOB = getName_N_DOB(wac_sheet)
    leg_names_DOB = getName_N_DOB(leg_sheet)


    related_value = []

    #return an object of leg_names, wac_naes, ... and store it in related _value
    for wac_values in wac_names_DOB:
        for leg_values in leg_names_DOB:
            dif_name = {}
            dif_DOF = {}
            related_name = {}

            if wac_values["name"].intersection(leg_values["name"]):
                related_name = wac_values["name"].intersection(leg_values["name"])

                dif_name = wac_values["name"].difference(leg_values['name'])
                dif_name2 = leg_values["name"].difference(wac_values["name"])
                dif_name.update(dif_name2)

                dif_DOF = wac_values["dateOfBirth"].difference(leg_values["dateOfBirth"])

                if len(dif_name) <= 0:
                    dif_name = {"No mismatch"}

                if len(dif_DOF) <= 0:
                    dif_DOF = {"No mismatch"}

                # add the result to the related value array
                related_value.append({
                    "leg_names": leg_values["name"],
                    "wac_names": wac_values["name"],
                    "related_names": related_name,
                    "dif_names": dif_name,
                    "Ug_DOB": leg_values["dateOfBirth"],
                    "Wasc_DOB": wac_values["dateOfBirth"],
                    "dif_DOF": dif_DOF,

                })

    #print(related_value)

    def sortDate(dateSet):
        #sort the arrary  and remove the symbols we assigned ealier
        dateSet = list(dateSet)

        dateSet.sort()
        #day month year
        i = 0
        for _date in dateSet:
            if _date[0] in "dmy":
                dateSet[i] = _date[1:]
                i += 1

        return dateSet
    #print(sortDate({'y2002', 'd12', 'm04'}))

    # populate sheet row 1 as header
    report_sheet.append(["Ug_name", "Wac_name","related names", "Diff name", "Leg_DOB", "Wac_DOF", "Diff_DOB"])

    # populate the rest of the sheet rows
    for value in related_value:
        leg_names = " ".join(value["leg_names"]).title()
        wac_names = " ".join(value["wac_names"]).title()
        related_names = " ".join(value["related_names"]).title()
        dif_names = " ".join(value['dif_names']).title()
        leg_DOB = "/".join(sortDate(value['Ug_DOB']))
        wac_DOB = "/".join(sortDate(value["Wasc_DOB"]))
        dif_DOB = "/".join(sortDate(value['dif_DOF']))

        # adding values into the sheet in a row
        report_sheet.append([leg_names, wac_names, related_names, dif_names, leg_DOB, wac_DOB, dif_DOB])

    report_workbook.save(filename=reportSavedWorkbook)


wasc_leg_name_report("dummy3.xlsx", "dummy4.xlsx", "report.xlsx")
